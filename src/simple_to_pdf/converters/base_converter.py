from abc import ABC, abstractmethod
from pathlib import Path
from PIL import Image
import io
import openpyxl
import logging

logger = logging.getLogger(__name__)

class BaseConverter(ABC):

    SUPPORTED_FORMATS = {
        "pdf": {".pdf"},
        "excel": {".xls", ".xlsx"},
        "word": {".doc", ".docx"},
        "image": {".jpg", ".jpeg", ".png"},
        "presentation": {".ppt", ".pptx"}
    }

    def __init__(self,*, chunk_size: int = 10):
        self.chunk_size = chunk_size

    @staticmethod
    def make_chunks(lst, n):

        """Splits list lst into chunks of n elements."""

        for i in range(0, len(lst), n):
            yield lst[i:i + n]
    
    @abstractmethod
    def convert_to_pdf(self,*, files: list[tuple[int, Path]]) -> list[tuple[int, bytes]]:

        """Abstract method that all classes must implement"""

        pass

    @classmethod
    def get_supported_formats(cls):
        all_exts = []
        for exts in cls.SUPPORTED_FORMATS.values():
            all_exts.extend(exts)
        return tuple(all_exts)

    def is_pdf_file(self,*, file_path: Path) -> bool:
        return file_path.suffix.lower() == self.SUPPORTED_FORMATS['pdf']
        
    def is_excel_file(self,*, file_path: Path) -> bool:
         return file_path.suffix.lower() in self.SUPPORTED_FORMATS['excel']
           
    def is_image_file(self,*, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.SUPPORTED_FORMATS['image']
    
    def is_word_file(self,*, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.SUPPORTED_FORMATS['word']
    
    def is_presentation_file(self,*, file_path: Path) -> bool:
        return file_path.suffix.lower() in self.SUPPORTED_FORMATS['presentation']
    
    from pathlib import Path

    def needs_conversion(self,*,file_path: Path) -> bool:

        """ Checks if the file extension is in SUPPORTED_FORMATS (excluding PDF)."""

        # Create a flat set of all extensions that need conversion
        # We skip 'pdf' during this process
        convertible_exts = {
            ext 
            for category, exts in self.SUPPORTED_FORMATS.items() 
            if category != "pdf" 
            for ext in exts
        }
        # Get the extension of the input file and check
        file_ext = file_path.suffix.lower()
        return file_ext in convertible_exts

    
    
    def convert_images_to_pdf(self,*, files: list[tuple[int,str]]) -> list[tuple[int, bytes]]:
        pdfs: list[tuple[int,bytes]] = []
        for idx,path_str in files:
            path=Path(path_str)
            if path.exists():
                try:
                    img = Image.open(path).convert("RGB")
                    buffer = io.BytesIO()
                    img.save(buffer,format = "PDF")
                    pdfs.append((idx,buffer.getvalue()))
                except Exception as e:
                     logger.error(f"⚠️ [{idx}] Error: failed to convert image {path} ({e})", exc_info = True)
            else:
                logger.warning(f"⚠️ [{idx}] Skipped: {path} (not an image or missing)")
        return pdfs
    
    def get_excel_width(self,*, file_path: Path) -> dict[Path,int]:
        
        workbook = openpyxl.load_workbook(filename = file_path, data_only = True)
        report: dict[str,int] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            width = sheet.max_column if sheet.max_column else 0
            report[sheet_name] = width
        workbook.close
        return report