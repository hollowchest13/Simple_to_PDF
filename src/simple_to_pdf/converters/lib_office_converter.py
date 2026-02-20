import logging
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl

from simple_to_pdf.converters.img_converter import ImageConverter
from simple_to_pdf.converters.models import ConversionResult
from simple_to_pdf.converters.libre_excel_formater import OpenPyXlFormatterMixin

logger = logging.getLogger(__name__)


class LibreOfficeConverter(ImageConverter,OpenPyXlFormatterMixin):
    SUPPORTED_FORMATS = {
        "table": {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".xlsb", ".ods", ".csv"},
        "document": {".doc", ".docx", ".odt", ".rtf", ".txt"},
        "presentation": {".ppt", ".pptx", ".odp"},
    }

    def __init__(self, *, soffice_path: str, chunk_size: int = 30):
        # Call constructor of base class, so it can initialize its data
        super().__init__(chunk_size=chunk_size)
        self.soffice_path = soffice_path
        self.SUPPORTED_FORMATS = self.get_supported_formats()

    def convert_to_pdf(
        self, *, files: list[tuple[int, Path]]
    ) -> ConversionResult:
        docs: list[tuple[int, Path]] = []
        imgs: list[tuple[int, Path]] = []
        final_result: ConversionResult = ConversionResult()

        for idx, path in files:
            if not path.exists():
                continue
            if (
                self.is_table_file(file_path=path)
                or self.is_document_file(file_path=path)
                or self.is_presentation_file(file_path=path)
            ):
                docs.append((idx, path))
            elif self.is_image_file(file_path=path):
                imgs.append((idx, path))
        docs_res = self._convert_docs_to_pdf(files=docs)
        img_res = self._convert_images_to_pdf(files=imgs)

        final_result.successful.extend(docs_res.successful)
        final_result.failed.extend(docs_res.failed)
        final_result.successful.extend(img_res.successful)
        final_result.failed.extend(img_res.failed)

        return final_result

    def _convert_docs_to_pdf(
        self, *, files: list[tuple[int, Path]]
    ) -> ConversionResult:
        all_results = ConversionResult()
        for chunk in self.make_chunks(files, self.chunk_size):
            # Call the new method that handles one chunk
            chunk_res = self._convert_chunk(chunk=chunk)
            all_results.successful.extend(chunk_res.successful)
            all_results.failed.extend(chunk_res.failed)
        return all_results

    def _run_libreoffice_format_conversion(
        self, *, input_paths: list[Path], out_dir: Path
    ):
        """all tables to xlsx conversion"""

        command = [
            self.soffice_path,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
        ] + [str(p) for p in input_paths]

        num_files = len(input_paths)
        timeout = 10 + num_files * 5

        try:
            subprocess.run(command, check=True, capture_output=True, timeout=timeout)
            # Deleting old xls
            for p in input_paths:
                if p.exists():
                    p.unlink()
        except subprocess.TimeoutExpired:
            logger.error(
                f"❌ LibreOffice timed out after {timeout} seconds for {num_files} files: {[p.name for p in input_paths]}"
            )
        except subprocess.CalledProcessError as e:
            logger.error(
                f"❌ LibreOffice conversion error for {[p.name for p in input_paths]}: {e.stderr.decode()}",
                exc_info=True,
            )
        except Exception as e:
            logger.error(
                f"❌ Unexpected error during table conversion: {e}", exc_info=True
            )

    def _prepare_temp_files(
        self, *, chunk: list[tuple[int, Path]], tmp_path: Path
    ) -> list[Path]:
        """Copy files in temporary directory with index prefix."""

        paths: list[Path] = []
        for idx, original_path in chunk:
            temp_name = f"{idx}_{original_path.name}"
            temp_file_path = tmp_path / temp_name
            shutil.copy2(original_path, temp_file_path)
            paths.append(temp_file_path)
        return paths

    def _convert_chunk(self, *, chunk: list[tuple[int, Path]]) -> ConversionResult:
        """Logic for processing one chunk of files."""

        to_convert_exts: list[str] = [".xls", ".xlsb", ".ods", ".csv"]
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)

            # Prepare files (copy with index prefix)
            all_tmp_paths: list[Path] = self._prepare_temp_files(
                chunk=chunk, tmp_path=tmp_path
            )
            # Selecting EVERYTHING that requires conversion to .xlsx before running openpyxl
            xls_to_convert: list[Path] = [
                p for p in all_tmp_paths if p.suffix.lower() in to_convert_exts
            ]
            if xls_to_convert:
                self._run_libreoffice_format_conversion(
                    input_paths=xls_to_convert, out_dir=tmp_path
                )
                all_tmp_paths = self._update_paths(
                    all_paths=all_tmp_paths, to_check_exts=to_convert_exts
                )
            for path in all_tmp_paths:
                if path.suffix.lower() == ".xlsx":
                    self._prepare_excel_scaling(file_path=path)
            input_paths = [str(p) for p in all_tmp_paths]

            # Run the conversion
            success = self._run_libreoffice_command(
                input_paths=input_paths, out_dir=tmp_path
            )

            # if command was successful, collect results
            chunk_res: ConversionResult = ConversionResult()  # створюємо завчасно
            if success:
                chunk_res = self._collect_results(chunk=chunk, tmp_path=tmp_path)
            return chunk_res

    def _update_paths(
        self, *, all_paths: list[Path], to_check_exts: list[str]
    ) -> list[Path]:
        updated = []
        for p in all_paths:
            if p.suffix.lower() in to_check_exts:
                expected = p.with_suffix(".xlsx")
                if expected.exists():
                    updated.append(expected)
                else:
                    logger.warning(
                        f"⚠️ Failed to xlsx conversion {p.name}, keeping as {p.suffix}"
                    )
                    updated.append(p)
            else:
                # If it's .xlsx or any other file - just adding it back
                updated.append(p)
        return updated

    def _run_libreoffice_command(
        self, *, input_paths: list[str], out_dir: Path
    ) -> bool:
        """Run the LibreOffice conversion command."""

        command = [
            self.soffice_path,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(out_dir),
        ] + input_paths
        num_files = len(input_paths)
        timeout = 10 + num_files * 5
        try:
            subprocess.run(command, check=True, capture_output=True, timeout=timeout)
            return True
        except subprocess.TimeoutExpired:
            logger.error(
                f"❌ LibreOffice timed out after {timeout} seconds for {num_files} files"
            )
            return False
        except subprocess.CalledProcessError as e:
            logger.error(f"❌ LibreOffice error: {e}", exc_info=True)
            return False
        except Exception as e:
            logger.error(
                f"❌ Unexpected error during table conversion: {e}", exc_info=True
            )
            return False

    def _collect_results(
        self, *, chunk: list[tuple[int, Path]], tmp_path: Path
    ) -> ConversionResult:
        """Reads created PDF files into memory."""
        res = ConversionResult()
        for idx, original_path in chunk:
            expected_pdf = tmp_path / f"{idx}_{original_path.stem}.pdf"
            if expected_pdf.exists():
                res.successful.append((idx, expected_pdf.read_bytes()))
            else:
                logger.warning(f"⚠️ Failed conversion to pdf: {expected_pdf.name}")
                res.failed.append((idx, original_path))
        return res

    def get_excel_width(self, *, file_path: Path) -> dict[str, int]:
        workbook = openpyxl.load_workbook(filename=file_path, data_only=True)
        report: dict[str, int] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            width = sheet.max_column if sheet.max_column else 0
            report[sheet_name] = width
        workbook.close()
        return report
